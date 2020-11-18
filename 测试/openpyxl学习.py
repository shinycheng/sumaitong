from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheeet")
ws.title = "New"
ws3 = wb["New"]
c = ws['A4']
ws['a4'] = 4
d = ws.cell(row=4,column=2,value=10)
for i in range(1,10):
    for y in range(1,101):
        ws.cell(row=i,column=y,value=y)
