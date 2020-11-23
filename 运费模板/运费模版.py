import os
from openpyxl import Workbook
from openpyxl import load_workbook
os.system("python 速卖通运费模板.py")
wb = load_workbook('速卖通运费模板.xlsx')
wb1 = load_workbook('20201019-727886255-242241330-AliExpress 无忧物流-标准.xlsx')
ws1 = wb1.get_sheet_by_name('线路编码')
wb2 = load_workbook('20201019-727886255-242241330-中国邮政挂号小包.xlsx')
ws2 = wb2.get_sheet_by_name('线路编码')
wb3 = load_workbook('20201019-727886255-242241330-燕文航空挂号小包.xlsx')
ws3 = wb3.get_sheet_by_name('线路编码')
wswuyou = wb.get_sheet_by_name('AliExpress无忧物流-标准')
wszhongyou = wb.get_sheet_by_name('中国邮政挂号小包')
wsyanwen = wb.get_sheet_by_name('燕文航空挂号小包')
row1 = wswuyou.max_row
col1 = wswuyou.max_column
row2 = wszhongyou.max_row
col2 = wszhongyou.max_column
row3 = wsyanwen.max_row
col3 = wsyanwen.max_column
#AliExpress无忧物流-标准
for row in range(4,ws1.max_row+1):
    for col in range(1,ws1.max_column+1):
        ws1.cell(row=row, column=col).value = None
for i in range(1,row1):
    ws1.cell(row=i+2,column=1).value = wswuyou.cell(row=i+1,column=1).value
    ws1.cell(row=i+2,column=5).value = ws1.cell(row=3,column=5).value
    ws1.cell(row=i+2,column=9).value = ws1.cell(row=3,column=9).value
    ws1.cell(row=i+2,column=10).value = ws1.cell(row=3,column=10).value
    ws1.cell(row=i+2,column=11).value = wswuyou.cell(row=i+1,column=col1).value
    if ws1.cell(row=i+2,column=11).value<0:
        ws1.cell(row=i + 2, column=11).value =0
    ws1.cell(row=i+2, column=12).value = ws1.cell(row=3, column=12).value
    ws1.cell(row=i+2, column=13).value = ws1.cell(row=3, column=13).value
wb1.save('20201019-727886255-242241330-AliExpress 无忧物流-标准.xlsx')
#中国邮政挂号小包
for row in range(4,ws2.max_row+1):
    for col in range(1,ws2.max_column+1):
        ws2.cell(row=row, column=col).value = None
for i in range(1,row2):
    ws2.cell(row=i+2,column=1).value = wszhongyou.cell(row=i+1,column=1).value
    ws2.cell(row=i+2,column=5).value = ws2.cell(row=3,column=5).value
    ws2.cell(row=i+2,column=9).value = ws2.cell(row=3,column=9).value
    ws2.cell(row=i+2,column=10).value = ws2.cell(row=3,column=10).value
    ws2.cell(row=i+2,column=11).value = wszhongyou.cell(row=i+1,column=col2).value
    if ws2.cell(row=i+2,column=11).value<0:
        ws2.cell(row=i + 2, column=11).value =0
    ws2.cell(row=i+2, column=12).value = ws2.cell(row=3, column=12).value
    ws2.cell(row=i+2, column=13).value = ws2.cell(row=3, column=13).value
wb2.save('20201019-727886255-242241330-中国邮政挂号小包.xlsx')
#燕文
for row in range(4,ws3.max_row+1):
    for col in range(1,ws3.max_column+1):
        ws3.cell(row=row, column=col).value = None
for i in range(1,row3):
    ws3.cell(row=i+2,column=1).value = wsyanwen.cell(row=i+1,column=1).value
    ws3.cell(row=i+2,column=5).value = ws3.cell(row=3,column=5).value
    ws3.cell(row=i+2,column=9).value = ws3.cell(row=3,column=9).value
    ws3.cell(row=i+2,column=10).value = ws3.cell(row=3,column=10).value
    ws3.cell(row=i+2,column=11).value = wsyanwen.cell(row=i+1,column=col3).value
    if ws3.cell(row=i+2,column=11).value<0:
        ws3.cell(row=i + 2, column=11).value =0
    ws3.cell(row=i+2, column=12).value = ws3.cell(row=3, column=12).value
    ws3.cell(row=i+2, column=13).value = ws3.cell(row=3, column=13).value
wb3.save('20201019-727886255-242241330-燕文航空挂号小包.xlsx')
