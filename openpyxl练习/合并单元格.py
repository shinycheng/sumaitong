from openpyxl.workbook import Workbook
wb = Workbook()
ws = wb.active
ws.merge_cells('A2:D2')
ws.unmerge_cells('A2:D2')
ws.merge_cells(start_row=2,start_column=1,end_row=4,end_column=4)
ws.unmerge_cells(start_row=2,start_column=1,end_row=4,end_column=4)
wb.save('c.xlsx')