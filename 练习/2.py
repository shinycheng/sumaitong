import requests
import json
import threading
import queue
from openpyxl import Workbook
from openpyxl import load_workbook
#cooike获取，去搜索名字core-indicators?dateType=recent30
wb = Workbook()
ws = wb.active
lowb = load_workbook('行业趋势代码.xlsx')
lows = lowb.active
aparams = []
for i in range(lows.max_row+1):
    js0 = lows.cell(row=i+1,column=2).value
    if js0 == 0:
        js0 = -9999
    js1 = lows.cell(row=i+1,column=3).value
    aparams.append((
        ('dateType', 'recent30'),
        ('dateRange', '2020-11-10|2020-12-09'),
        ('country', 'ALL'),
        ('platform', 'ALL'),
        ('cateId', js1),
        ('parentCateId', js0),))
for param in aparams:
    params = param
    cookies = {
        'ali_apache_id': '11.180.122.25.1590413831636.233412.0',
        'cna': 'CLhSFwH4MmcCATyxVFFC+YKD',
        '_ga': 'GA1.2.57210478.1590413834',
        'aep_common_f': 'mDtIp9CrgzENqa+IZhfyCtiouYyIw14Pd4LcNigtdzFzLaboM1Qcng==',
        'ae-msite-city': '',
        'ae-msite-province': '',
        '_lang': 'zh_CN',
        '_fbp': 'fb.1.1604130214911.337659290',
        '_hvn_login': '13',
        'aep_usuc_t': 'ber_l=A0',
        '_gid': 'GA1.2.1544221239.1607224526',
        '_tb_token_': '51DTQmx65H4ahPJRE19W_CxG',
        'xlly_s': '1',
        '_m_h5_tk': '1e9f41afd8d118121ea62fa37d30c88b_1607355828831',
        '_m_h5_tk_enc': '3f9bcb30915f6c6f6aaf4fece99fddbd',
        'acs_usuc_t': 'acs_rt=7bdc0d6a49984953b4f812dde53f9289&x_csrf=l99xflgdlg74',
        'havana_tgc': 'eyJjcmVhdGVUaW1lIjoxNjA3MzQ4Mjc5MzY0LCJsYW5nIjoiemhfQ04iLCJwYXRpYWxUZ2MiOnsiYWNjSW5mb3MiOnsiMTMiOnsiYWNjZXNzVHlwZSI6MSwibWVtYmVySWQiOjIyMDgwOTYxMzgyMTYsInRndElkIjoiMW55bFRwVl9USnp4V0I4QUlRODBXUWcifX19fQ',
        'xman_us_t': 'x_lid=cn1532108722sapg&sign=y&rmb_pp=yolocity_001@163.com&x_user=+tuCa+dNu2Spq50ApaVDIXOE8tQg/F+q+QCIhLmhH9k=&ctoken=w9u_sykgyhks&need_popup=y&l_source=aliexpress',
        'xman_f': '5QMP64SC0bPGGzMbXZyD/UYHssiy5HQ3dAwnkE9Z02JkaD1uTkj4UrgyfdAfXhifmdV+Q4VQvdBrTR2lFgUCAlGTaNZ2bBx3D3S/Ea0YMW6Jvlv/TAmEWeUYLxecqvwkRmRBMYYnQzlN+txG+IorHl++yH4IENMthGehfXK0mAH5O9UWDmap5SasBgThXWIXbvs0jLHVdwYwgofPosBKxF7jgBezabNZ3cl/cMP2p6+NTNu/550N+DG7AA3IORKXtkBMigMnWizW4gQ9M06W9UjJfJC98aIuk8KIqGH35K7hvC7q2O9WdSXz5k7AfSpWBrDlBQ72+JgHu2tqvewsTXR+nUAjHmJG/Oc0MnjnoUw82w3nrolp0rZDstSe/2Zg98LXaOmSXQHXOCdmB00uiCA18Nk93gHrL/zuHamDozsNw9NXMWTgMQo9TaAy9xAd',
        'aep_usuc_f': 'isfm=y&site=glo&province=917477670000000000&city=917477679976000000&c_tp=USD&x_alimid=242241330&iss=y&s_locale=zh_CN&region=RU&b_locale=en_US',
        'ali_apache_tracktmp': 'W_signed=Y',
        'ali_apache_track': 'mt=2|ms=|mid=cn1532108722sapg',
        'intl_locale': 'en_US',
        'aep_history': 'keywords%5E%0Akeywords%09%0A%0Aproduct_selloffer%5E%0Aproduct_selloffer%091005001667875263%091005001617950091%094001267147903%091005001667875263%091005001795232827%091005001617950091%091005001759718226%091005001617950091',
        'intl_common_forever': 'Ui2xubitZtkOICjDwJhprqw+N0uYQNTfLhAkEazU8b2fIIJ/Jb3pNA==',
        'xman_us_f': 'zero_order=y&x_locale=en_US&x_l=1&last_popup_time=1590495451952&x_user=CN|null|null|cnfm|242241330&no_popup_today=n&x_lid=cn1532108722sapg&x_c_chg=0&x_c_synced=0&x_as_i=%7B%22cookieCacheEffectTime%22%3A1607354759029%2C%22isCookieCache%22%3A%22Y%22%2C%22ms%22%3A%220%22%7D&acs_rt=a4d1fc34026042d7b6480fcd59a2ba5c',
        'xman_t':'OX+1ISffSDHPla7Og0Xjz0k1y4jVsIS8o+cfBSwhOLliUwqPD8LU/iAp31/ORjOBhowAi9yY2itG6VAhNYwqFwhNKN4kbnL8sBk9hUAo8yaOjvvrYX+e6N3SMsJnKlWYLke5kuO8DT2zxtjBDtmuz3ZNQ9NWBsRfx7OfvOKNv4zbC41cTd700GPtuLgOBWiaMc/ol1OI5g/I/FSmoVDmocMgNNrrSs13Pw6Topdx9mK/ur/ISBIWcp65arfMLmhH5Ne0u2mY2RNrHHomZ/l3oAEEe13s7BYZt9QkTNQ2sGJEzTutOqNpLFoVftJrbNpOtkzLXIsBlstC13ualBo67aZb8JyqZkYmno/4N8+HBk2t1JS/R8v9OIOucWj1T2qzHR8GQ/eAaZopafnAPUHhjuACeIN8zQpxFBufCCqJcamHUu8EZuJRWcCiGA13DhtnSmptVIgzvEYHB4sSqLNaQghjtDu/1SrMvaraQRDhqNktzAV7s/pG/l8ExeJM7UGlPeev16JdDMDHWr0xGsjwoNo8Yh4jXARg8sgGWsGfThZPC5Y649U0w/jvXO3rNTyzrdqa0YWYb/lKRXs+g7nfjbYLG2PPja29iG/3hPiB038UgIvoSV15u9W9L1N0O/4K5+WXyd+biCi+6G7bbTRGctvf0OlkDByF',
        'l': 'eBO7cchmQeDU6-xvBO5aKurza77TVBOjCcVzaNbMiIncC6fRmPJpFqKQKtOooUxRJ8XATzYH49fRGnyODX-u7bLMo2_TPWv3dok2B',
        'isg': 'BM_PLqzwzAjnsslxGwRoD9vjXmXZ9CMWX1NJ2-HbQDxcsPqy68TsZj5ms-AO-PuO',
        'tfstk': 'ccD5ByaX82H4KHOFU0t4YgdG2UzcakX703amVbjbj3RdzicYks01QPssVDbAH5Ef.',
    }
    headers = {
        'authority': 'sycm.aliexpress.com',
        'sec-ch-ua': '"Chromium";v="86", "\\"Not\\\\A;Brand";v="99", "Google Chrome";v="86"',
        'sec-ch-ua-mobile': '?0',
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 11_0_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36',
        'accept': '*/*',
        'sec-fetch-site': 'same-origin',
        'sec-fetch-mode': 'cors',
        'sec-fetch-dest': 'empty',
        'referer': 'https://sycm.aliexpress.com/mc/MarketOverview?spm=5261.newworkbench.aenewheader.7.7fe24edf5M0J8S&dateRange=2020-11-09%7C2020-12-08&dateType=recent30&cate=%7B%22parentId%22%3A18%2C%22cateId%22%3A201445239%2C%22cateName%22%3A%22%E8%88%9E%E8%B9%88%E8%BF%90%E5%8A%A8%22%2C%22cateFlag%22%3A2%7D&country=ALL&platform=ALL',
        'accept-language': 'zh-CN,zh;q=0.9,zh-TW;q=0.8,en;q=0.7,ru;q=0.6',
    }
    response = requests.get('https://sycm.aliexpress.com/api/market-dashboard/core-indicators', headers=headers,params=params, cookies=cookies)
    js = json.loads(response.text)
    data = js['data']
    ali_dict = {}
    ali_dict['父类金额占比'] = str((data['payAmtParentCatePercent']['value']) * 100) + '%'
    ali_dict['商品加购人数'] = data['itemAddCartBuyerCnt']['value']
    ali_dict['搜索指数'] = data['seUvIndex']['value']
    ali_dict['交易指数'] = data['noriskPayAmtIndex']['value']
    ali_dict['浏览商品数'] = data['visitedItemCnt']['value']
    ali_dict['商品收藏人数'] = data['wishlistBuyerCnt']['value']
    ali_dict['访客指数'] = data['uvIndex']['value']
    ali_dict['供需指数'] = data['supplyDemandIndex']['value']
    ali_dict['客单价'] = data['payPerBuyerAmt']['value']
    ali_dict['客单价'] = str((data['vstItemPercent']['value']) * 100) + '%'
    print(ali_dict)