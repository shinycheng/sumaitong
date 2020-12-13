from selenium import webdriver
import time
from openpyxl import load_workbook
#订单号路径
wb =load_workbook('/Users/zhanghang/Downloads/H76c80fd4ddea43338a089fcee5d00c5dX.xlsx')
ws = wb.get_sheet_by_name('1')
row_max = ws.max_row
driver = webdriver.Chrome()
driver.get("https://gsp.aliexpress.com/")
time.sleep(1.5)
#登陆
iframe = driver.switch_to.frame(driver.find_element_by_xpath('//*[@id="alibaba-login-box"]'))
name = driver.find_element_by_xpath('//*[@id="fm-login-id"]').send_keys('yolocity_001@163.com')
passsword = driver.find_element_by_xpath('//*[@id="fm-login-password"]').send_keys('Yolocity817')
driver.find_element_by_xpath('//*[@id="fm-login-submit"]').click()
###
#点击交易
time.sleep(5)
driver.find_element_by_xpath('//*[@id="dada-wrapper"]/div[1]/div[2]/div[1]/div[3]/div/div[3]/a/div/div[1]').click()
ha = driver.current_window_handle
#输入订单号
def dingdan():
    time.sleep(3)
    for i in range(0,row_max):
        dingdanhao = ws.cell(row = i+1,column = 1).value
        #订单号框
        time.sleep(2)
        driver.find_element_by_xpath('//*[@id="root"]/div/div/div/div/div/div/div[2]/div/div[2]/div/div[2]/div/div/div[1]/div/div/div/div[1]/div/span/input').send_keys(str(dingdanhao))
        time.sleep(2)
        #点击搜索框
        driver.find_element_by_xpath('//*[@id="root"]/div/div/div/div/div/div/div[2]/div/div[2]/div/div[2]/div/div/div[2]/div/button/i').click()
        time.sleep(2)
        #点击与买家联系
        driver.find_element_by_xpath('//*[@id="root"]/div/div/div/div/div/div/div[2]/div/div[5]/div/div/table/tbody/tr/td[2]/div/div/div/div[2]/a').click()
        time.sleep(2)
        for handle in driver.window_handles:
            if handle != ha:
                driver.switch_to.window(handle)
                time.sleep(0.5)
                pr = driver.find_element_by_xpath('//*[@id="_lzd-im-container"]/div/div/div[2]/div[1]/div[2]/div[1]/div[3]/div[2]/div[2]/textarea')
                pr.send_keys("Dear friend.\nThank you for shopping at our store\nYour shopping experience is very important to us and our business.\nIf you are satisfied with our product, we would like to invite you to leave positive feedback on our products.\nIf there is any quality or other problems, please don't give us a negative review immediately. Just contact me here and I'll provide you a good solution and solve it in time.\nThanks for your support. Wish you have a good day!\nRegards")
                driver.find_element_by_xpath('//*[@id="_lzd-im-container"]/div/div/div[2]/div[1]/div[2]/div[1]/div[3]/div[2]/div[2]/div/button').click()
                time.sleep(0.5)
                driver.close()
                driver.switch_to.window(ha)
                driver.find_element_by_xpath('//*[@id="dada-wrapper"]/div[1]/div[2]/div[1]/div[3]/div/div[3]/a/div/div[1]').click()
                time.sleep(2)
dingdan()
