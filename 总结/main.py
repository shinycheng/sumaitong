from config import *
import os
from openpyxl import load_workbook
from openpyxl import Workbook
os.system("python 速卖通运费模板.py")
class Sumaitong:
    #制作sku
    def sku():
        for i in range(len(color)):
            for j in range(len(size)):
                try:
                    sku_list.append(name+'-'+str(Color[color[i]])+'-'+size[j])
                except:
                    print(color[i]+'不再Color列表里')
                    break
    #28个国家折前价
    def countrys_28():
        # 标题
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = '尺码'
        ws.cell(row=1, column=2).value = '成本'
        ws.cell(row=1, column=3).value = '重量'
        ws.cell(row=1, column=4).value = '毛利'
        ws.cell(row=1, column=5).value = '折扣'
        ws.cell(row=1, column=6).value = '零售价'
        ws.cell(row=1, column=7).value = '促销价'
        ws.cell(row=1, column=8).value = '库存'
        ws.cell(row=1, column=9).value = 'sku'
        ws.cell(row=1, column=10).value = '调价方式'
        ws.cell(row=1, column=11).value = '属性对'
        for j in range(0, 28):
            ws.cell(row=1, column=j + 12).value = countrys[j]
        # 结果
        for x in range(len(sku_list)):
            ws.cell(row=x + 2, column=9).value = sku_list[x]
            for y in range(len(color)):
                max_i = len(size)
                for i in range(len(size)):
                    ws.cell(row=i + 2+ max_i*y, column=1).value = size[i]
                    ws.cell(row=i + 2+ max_i*y, column=2).value = product_cost[i]
                    ws.cell(row=i + 2+ max_i*y, column=3).value = product_weight[i]
                    ws.cell(row=i + 2+ max_i*y, column=4).value = maoli
                    ws.cell(row=i + 2+ max_i*y, column=5).value = zhekou
                    ws.cell(row=i + 2 + max_i * y, column=6).value = product_sale_price/zhekou
                    ws.cell(row=i + 2 + max_i * y, column=7).value = None
                    ws.cell(row=i + 2 + max_i * y, column=8).value = 99
                    ws.cell(row=i + 2 + max_i * y, column=10).value = 'absolute'
                    ws.cell(row=i + 2 + max_i * y, column=11).value = None
                    for j in range(0, 28):
                        ws.cell(row=i + 2+ max_i*y, column=j + 12).value = ((feilv[j] * product_weight[i] + guahaofei[j] + product_cost[i]) / ((shouxufei - maoli) * USD_to_RMB)) / zhekou
        wb.save(name + '分析.xlsx')
    #运费
    def yunfei():
        wb = load_workbook('速卖通运费模板.xlsx')
        wb1 = load_workbook('20201019-727886255-242241330-AliExpress 无忧物流-标准.xlsx')
        ws1 = wb1.get_sheet_by_name('线路编码')
        wb2 = load_workbook('20201019-727886255-242241330-中国邮政挂号小包.xlsx')
        ws2 = wb2.get_sheet_by_name('线路编码')
        wb3 = load_workbook('20201019-727886255-242241330-燕文航空挂号小包.xlsx')
        ws3 = wb3.get_sheet_by_name('线路编码')
        wswuyou = wb.get_sheet_by_name('AliExpress无忧物流-标准')
        wszhongyou = wb.get_sheet_by_name('中国邮政挂号小包')
        wsyanwen = wb.get_sheet_by_name('燕文航空挂号小包')
        row1 = wswuyou.max_row
        col1 = wswuyou.max_column
        row2 = wszhongyou.max_row
        col2 = wszhongyou.max_column
        row3 = wsyanwen.max_row
        col3 = wsyanwen.max_column
        # AliExpress无忧物流-标准
        for row in range(4, ws1.max_row + 1):
            for col in range(1, ws1.max_column + 1):
                ws1.cell(row=row, column=col).value = None
        for i in range(1, row1):
            ws1.cell(row=i + 2, column=1).value = wswuyou.cell(row=i + 1, column=1).value
            ws1.cell(row=i + 2, column=5).value = ws1.cell(row=3, column=5).value
            ws1.cell(row=i + 2, column=9).value = ws1.cell(row=3, column=9).value
            ws1.cell(row=i + 2, column=10).value = ws1.cell(row=3, column=10).value
            ws1.cell(row=i + 2, column=11).value = wswuyou.cell(row=i + 1, column=col1).value
            if ws1.cell(row=i + 2, column=11).value < 0:
                ws1.cell(row=i + 2, column=11).value = 0
            ws1.cell(row=i + 2, column=12).value = ws1.cell(row=3, column=12).value
            ws1.cell(row=i + 2, column=13).value = ws1.cell(row=3, column=13).value
        wb1.save('20201019-727886255-242241330-AliExpress 无忧物流-标准.xlsx')
        # 中国邮政挂号小包
        for row in range(4, ws2.max_row + 1):
            for col in range(1, ws2.max_column + 1):
                ws2.cell(row=row, column=col).value = None
        for i in range(1, row2):
            ws2.cell(row=i + 2, column=1).value = wszhongyou.cell(row=i + 1, column=1).value
            ws2.cell(row=i + 2, column=5).value = ws2.cell(row=3, column=5).value
            ws2.cell(row=i + 2, column=9).value = ws2.cell(row=3, column=9).value
            ws2.cell(row=i + 2, column=10).value = ws2.cell(row=3, column=10).value
            ws2.cell(row=i + 2, column=11).value = wszhongyou.cell(row=i + 1, column=col2).value
            if ws2.cell(row=i + 2, column=11).value < 0:
                ws2.cell(row=i + 2, column=11).value = 0
            ws2.cell(row=i + 2, column=12).value = ws2.cell(row=3, column=12).value
            ws2.cell(row=i + 2, column=13).value = ws2.cell(row=3, column=13).value
        wb2.save('20201019-727886255-242241330-中国邮政挂号小包.xlsx')
        # 燕文
        for row in range(4, ws3.max_row + 1):
            for col in range(1, ws3.max_column + 1):
                ws3.cell(row=row, column=col).value = None
        for i in range(1, row3):
            ws3.cell(row=i + 2, column=1).value = wsyanwen.cell(row=i + 1, column=1).value
            ws3.cell(row=i + 2, column=5).value = ws3.cell(row=3, column=5).value
            ws3.cell(row=i + 2, column=9).value = ws3.cell(row=3, column=9).value
            ws3.cell(row=i + 2, column=10).value = ws3.cell(row=3, column=10).value
            ws3.cell(row=i + 2, column=11).value = wsyanwen.cell(row=i + 1, column=col3).value
            if ws3.cell(row=i + 2, column=11).value < 0:
                ws3.cell(row=i + 2, column=11).value = 0
            ws3.cell(row=i + 2, column=12).value = ws3.cell(row=3, column=12).value
            ws3.cell(row=i + 2, column=13).value = ws3.cell(row=3, column=13).value
        wb3.save('20201019-727886255-242241330-燕文航空挂号小包.xlsx')



if __name__ == '__main__':
    Sumaitong.sku()
    Sumaitong.countrys_28()
    Sumaitong.yunfei()
